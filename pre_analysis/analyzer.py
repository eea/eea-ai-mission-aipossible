"""Run pre-analysis on Excel data sources using AI."""

import json
import time
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

from env_settings import get_bool_setting
from pre_analysis.utils import normalize_text, output_path_for_row, parse_pattern, truncate_text


def _load_sheet(input_file: Path, sheet_name: str):
    workbook = load_workbook(filename=input_file, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    return workbook[sheet_name]


def _find_column_index(worksheet, header_row: int, column_header: str) -> int:
    target = normalize_text(column_header)
    for cell in worksheet[header_row]:
        if normalize_text(cell.value) == target:
            return int(cell.column)
    raise ValueError(f"Column header not found: {column_header}")


def iter_column_rows(
    input_file: Path,
    sheet_name: str,
    column_header: str,
    header_row: int,
    start_row: int | None,
    max_rows: int | None,
) -> Iterator[tuple[int, str]]:
    worksheet = _load_sheet(input_file, sheet_name)
    column_index = _find_column_index(worksheet, header_row, column_header)
    row_start = start_row or (header_row + 1)
    count = 0
    for row in worksheet.iter_rows(
        min_row=row_start,
        min_col=column_index,
        max_col=column_index,
    ):
        cell = row[0]
        yield cell.row, cell.value
        count += 1
        if max_rows is not None and count >= max_rows:
            break


def _build_row_key(
    input_file: Path,
    sheet_name: str,
    column_header: str,
    row_index: int,
) -> str:
    return f"{input_file}|{sheet_name}|{column_header}|{row_index}"


def should_skip(output_path: Path) -> bool:
    """Return True when output already exists and is non-empty."""
    return output_path.exists() and output_path.stat().st_size > 0


def analyze_row(
    input_file: Path,
    sheet_name: str,
    column_header: str,
    header_row: int,
    row_index: int,
    raw_value: str | None,
    output_dir: Path,
    client,
    max_chars: int,
    overwrite: bool = False,
) -> tuple[Path, dict]:
    """Analyze a single row and write output JSON."""
    key = _build_row_key(input_file, sheet_name, column_header, row_index)
    output_path = output_path_for_row(output_dir, key)
    if not overwrite and should_skip(output_path):
        return output_path, {"processed": False, "skipped": True, "skip_reason": "exists"}

    normalized = normalize_text(raw_value)
    truncated_text, truncated, original_length = truncate_text(normalized, max_chars)
    ensure_ascii = get_bool_setting("JSON_ENSURE_ASCII", default=False)

    if not truncated_text:
        now = datetime.now(timezone.utc)
        payload = {
            "source": {
                "input_file": str(input_file),
                "sheet_name": sheet_name,
                "column_header": column_header,
                "header_row": header_row,
                "row_index": row_index,
            },
            "content": {
                "raw": "" if raw_value is None else str(raw_value),
                "normalized": truncated_text,
                "truncated": truncated,
                "original_length": original_length,
                "max_chars": max_chars,
            },
            "ai": {
                "provider": getattr(client, "provider", "") if client else "",
                "model": getattr(client, "model", "stub") if client else "stub",
                "prompt_version": getattr(client, "prompt_version", "") if client else "",
                "result_raw": "",
            },
            "timing": {
                "started_at": now.isoformat(),
                "completed_at": now.isoformat(),
                "elapsed_seconds": 0.0,
            },
            "status": {
                "processed": False,
                "skipped": True,
                "skip_reason": "empty",
            },
        }
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(payload, ensure_ascii=ensure_ascii, indent=2), encoding="utf-8")
        return output_path, payload["status"]

    analysis_started_at = datetime.now(timezone.utc)
    analysis_start_perf = time.perf_counter()
    ai_payload = client.analyze(truncated_text) if client else {}
    analysis_completed_at = datetime.now(timezone.utc)
    analysis_elapsed_seconds = time.perf_counter() - analysis_start_perf

    raw_result = ai_payload.get("result") or ""
    parsed = parse_pattern(raw_result) or {}

    payload = {
        "source": {
            "input_file": str(input_file),
            "sheet_name": sheet_name,
            "column_header": column_header,
            "header_row": header_row,
            "row_index": row_index,
        },
        "content": {
            "raw": "" if raw_value is None else str(raw_value),
            "normalized": truncated_text,
            "truncated": truncated,
            "original_length": original_length,
            "max_chars": max_chars,
        },
        "ai": {
            "provider": ai_payload.get("provider") or "",
            "model": getattr(client, "model", "stub") if client else "stub",
            "prompt_version": ai_payload.get("prompt_version") or "",
            "system_prompt": ai_payload.get("system_prompt") or "",
            "user_prompt": ai_payload.get("user_prompt") or "",
            "result_raw": raw_result,
            "pattern": parsed.get("pattern", ""),
            "explanation": parsed.get("explanation", ""),
            "confidence": parsed.get("confidence", None),
        },
        "timing": {
            "started_at": analysis_started_at.isoformat(),
            "completed_at": analysis_completed_at.isoformat(),
            "elapsed_seconds": analysis_elapsed_seconds,
        },
        "status": {
            "processed": True,
            "skipped": False,
            "skip_reason": "",
        },
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=ensure_ascii, indent=2), encoding="utf-8")
    return output_path, payload["status"]


def run_batch(
    input_file: Path,
    sheet_name: str,
    column_header: str,
    header_row: int,
    start_row: int | None,
    max_rows: int | None,
    output_dir: Path,
    client,
    max_chars: int,
    verbose: bool = True,
    overwrite: bool = False,
    dry_run: bool = False,
) -> dict:
    """Run pre-analysis over a column with minimal console output."""
    rows_dir = output_dir / "rows"
    count = 0
    skipped: Counter[str] = Counter()
    processed_rows = []

    for row_index, raw_value in iter_column_rows(
        input_file,
        sheet_name,
        column_header,
        header_row,
        start_row,
        max_rows,
    ):
        key = _build_row_key(input_file, sheet_name, column_header, row_index)
        output_path = output_path_for_row(rows_dir, key)
        if dry_run:
            if not overwrite and should_skip(output_path):
                skipped["exists"] += 1
                if verbose:
                    print(f"skip: {output_path.name}")
            else:
                if verbose:
                    print(f"would save: {output_path.name}")
            count += 1
            continue

        output_path, status = analyze_row(
            input_file=input_file,
            sheet_name=sheet_name,
            column_header=column_header,
            header_row=header_row,
            row_index=row_index,
            raw_value=raw_value,
            output_dir=rows_dir,
            client=client,
            max_chars=max_chars,
            overwrite=overwrite,
        )
        processed_rows.append(
            {
                "row_index": row_index,
                "output_path": output_path,
                "status": status,
            }
        )
        if status.get("skipped"):
            skipped[status.get("skip_reason", "unknown")] += 1
            if verbose:
                print(f"skip: {output_path.name}")
        elif verbose:
            print(f"saved: {output_path.name}")
        count += 1

    if verbose:
        print(f"done: processed={count} skipped={sum(skipped.values())}")

    return {
        "processed": count,
        "skipped": dict(skipped),
        "rows": processed_rows,
    }
