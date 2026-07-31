"""Module for analyzing Excel rows using AI and outputting structured results."""

import hashlib
import json
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

from analysis.analyzer import BatchItemResult, BatchRunStats
from analysis.utils import normalize_text, parse_answers
from env_settings import get_bool_setting


def _load_sheet(input_file: Path, sheet_name: str):
    workbook = load_workbook(filename=input_file, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    return workbook[sheet_name]


def _find_column_index(worksheet, header_row: int, column_name: str) -> int:
    target = normalize_text(column_name)
    for cell in worksheet[header_row]:
        if normalize_text(cell.value) == target:
            return int(cell.column)
    raise ValueError(f"Column header not found: {column_name}")


def iter_column_rows(
    input_file: Path,
    sheet_name: str,
    column_name: str,
    header_row: int,
    identifier_column: str | None = None,
) -> Iterator[tuple[int, str | None, str | None]]:
    worksheet = _load_sheet(input_file, sheet_name)
    column_index = _find_column_index(worksheet, header_row, column_name)
    identifier_index = _find_column_index(worksheet, header_row, identifier_column) if identifier_column else None
    if identifier_index is None:
        for row in worksheet.iter_rows(
            min_row=header_row + 1,
            min_col=column_index,
            max_col=column_index,
        ):
            cell = row[0]
            yield cell.row, cell.value, None
    else:
        min_col = min(column_index, identifier_index)
        max_col = max(column_index, identifier_index)
        for row in worksheet.iter_rows(min_row=header_row + 1, min_col=min_col, max_col=max_col):
            target_cell = row[column_index - min_col]
            id_cell = row[identifier_index - min_col]
            yield target_cell.row, target_cell.value, id_cell.value


def _output_path_for_row(output_dir: Path, row_key: str) -> Path:
    filename = hashlib.sha256(row_key.encode("utf-8")).hexdigest() + ".json"
    return output_dir / filename


def should_skip(output_path: Path) -> bool:
    """Return True when analysis output already exists and is non-empty."""
    return output_path.exists() and output_path.stat().st_size > 0


def analyze_row(
    input_file: Path,
    sheet_name: str,
    column_name: str,
    row_index: int,
    raw_value: str | None,
    output_dir: Path,
    client,
    use_case: str,
    source_type: str,
    source_path: str,
    overwrite: bool = False,
    row_identifier_column: str | None = None,
    row_identifier_value: str | None = None,
) -> tuple[Path, bool, float | None]:
    """Analyze a single Excel row and write output JSON."""
    key = f"{use_case}|{input_file}|{sheet_name}|{column_name}|{row_index}"
    output_path = _output_path_for_row(output_dir, key)
    if not overwrite and should_skip(output_path):
        return output_path, False, None

    text_value = normalize_text(raw_value)
    title = f"{Path(input_file).name} row {row_index}"
    url = f"excel://{Path(input_file).name}/{sheet_name}/{column_name}/{row_index}"
    analysis_started_at = datetime.now(timezone.utc)
    analysis_start_perf = time.perf_counter()
    ai_payload = client.analyze(text_value) if client and text_value else {}
    analysis_completed_at = datetime.now(timezone.utc)
    analysis_elapsed_seconds = time.perf_counter() - analysis_start_perf if text_value else 0.0

    raw_result = ai_payload.get("result") or ""
    answers = parse_answers(raw_result)

    analysis = {
        "url": url,
        "title": title,
        "summary": text_value,
        "provider": ai_payload.get("provider"),
        "model": getattr(client, "model", "stub"),
        "analyzed_at": analysis_completed_at.isoformat(),
        "analysis_started_at": analysis_started_at.isoformat(),
        "analysis_completed_at": analysis_completed_at.isoformat(),
        "analysis_elapsed_seconds": analysis_elapsed_seconds,
        "prompt": ai_payload.get("user_prompt", text_value),
        "system_prompt": ai_payload.get("system_prompt"),
        "prompt_version": ai_payload.get("prompt_version"),
        "ai_result": raw_result,
        "answers": answers,
        "source": {
            "use_case": use_case,
            "source_type": source_type,
            "source_path": source_path,
            "sheet_name": sheet_name,
            "column_name": column_name,
            "row_index": row_index,
            "row_identifier_column": row_identifier_column,
            "row_identifier": row_identifier_value,
        },
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    ensure_ascii = get_bool_setting("JSON_ENSURE_ASCII", default=False)
    output_path.write_text(json.dumps(analysis, ensure_ascii=ensure_ascii, indent=2), encoding="utf-8")
    return output_path, True, analysis_elapsed_seconds


def run_batch(  # noqa: C901
    input_file: Path,
    sheet_name: str,
    column_name: str,
    header_row: int,
    output_dir: Path,
    client,
    max_items: int | None = None,
    verbose: bool = True,
    overwrite: bool = False,
    dry_run: bool = False,
    use_case: str = "",
    source_type: str = "excel",
    source_path: str = "",
    row_identifier_column: str | None = None,
) -> BatchRunStats:
    """Run analysis over an Excel column and emit one JSON per row."""
    count = 0
    skipped = 0
    total_elapsed_seconds = 0.0
    items: list[BatchItemResult] = []

    for row_index, raw_value, identifier_value in iter_column_rows(
        input_file=input_file,
        sheet_name=sheet_name,
        column_name=column_name,
        header_row=header_row,
        identifier_column=row_identifier_column,
    ):
        row_key = f"{use_case}|{input_file}|{sheet_name}|{column_name}|{row_index}"
        output_path = _output_path_for_row(output_dir, row_key)
        url = f"excel://{Path(input_file).name}/{sheet_name}/{column_name}/{row_index}"
        if dry_run:
            saved = False
            if not overwrite and should_skip(output_path):
                skipped += 1
                if verbose:
                    print(f"skip: {output_path.name}")
            elif verbose:
                print(f"would save: {output_path.name}")
            items.append(
                BatchItemResult(
                    page_path=f"{input_file}#{sheet_name}:{column_name}:{row_index}",
                    output_path=str(output_path),
                    url=url,
                    saved=saved,
                    elapsed_seconds=None,
                )
            )
            count += 1
            if max_items and count >= max_items:
                break
            continue

        output_path, saved, elapsed = analyze_row(
            input_file=input_file,
            sheet_name=sheet_name,
            column_name=column_name,
            row_index=row_index,
            raw_value=raw_value,
            output_dir=output_dir,
            client=client,
            use_case=use_case,
            source_type=source_type,
            source_path=source_path,
            overwrite=overwrite,
            row_identifier_column=row_identifier_column,
            row_identifier_value=str(identifier_value) if identifier_value is not None else None,
        )
        if not saved:
            skipped += 1
            if verbose:
                print(f"skip: {output_path.name}")
        elif verbose:
            print(f"saved: {output_path.name}")

        if elapsed is not None:
            total_elapsed_seconds += elapsed
        items.append(
            BatchItemResult(
                page_path=f"{input_file}#{sheet_name}:{column_name}:{row_index}",
                output_path=str(output_path),
                url=url,
                saved=saved,
                elapsed_seconds=elapsed,
            )
        )
        count += 1
        if max_items and count >= max_items:
            break

    if verbose:
        total_minutes = total_elapsed_seconds / 60.0
        print(
            "done: "
            f"processed={count} skipped={skipped} "
            f"total_elapsed_seconds={total_elapsed_seconds:.2f} "
            f"total_elapsed_minutes={total_minutes:.2f}"
        )

    return BatchRunStats(
        processed=count,
        skipped=skipped,
        total_elapsed_seconds=total_elapsed_seconds,
        items=items,
    )
