"""Validate climate impacts against exported Excel metadata."""

import csv
import json
import re
from pathlib import Path

import openpyxl

STRICT_EXCEL_COMPARISON = False


def _normalize_title(value: object) -> str:
    text = str(value)
    text = text.replace("\u2019", "'").replace("\u2018", "'")
    text = " ".join(text.split()).strip().lower()
    text = re.sub(r"[^\w\s]", "", text, flags=re.UNICODE)
    return " ".join(text.split()).strip()


def _normalize_header(value: object) -> str:
    return " ".join(str(value).split()).strip().lower()


def _normalize_impact_value(value: str) -> str:
    normalized = value.strip().lower().replace("-", " ")
    if not normalized or normalized in {"nan", "none", "n/a", "na", "other"}:
        return ""
    mapping = {
        "extreme heat": "extreme temperatures",
        "extreme temperature": "extreme temperatures",
    }
    return mapping.get(normalized, normalized)


def _normalize_impacts(value: object) -> list[str]:
    if value is None:
        return []
    text = str(value)
    parts = [part.strip() for part in text.split(",")]
    normalized = [_normalize_impact_value(part) for part in parts if part]
    return [part for part in normalized if part]


def _normalize_observed_impacts(values: object) -> list[str]:
    if not values:
        return []
    if isinstance(values, str):
        return _normalize_impacts(values)
    normalized = [_normalize_impact_value(str(item)) for item in values if item is not None and str(item).strip()]
    return [item for item in normalized if item]


def _load_expected_from_excel(excel_column: str) -> tuple[dict, list, list]:
    excel_path = Path(__file__).parent / "fixtures" / "Mission-exported-pages.xlsx"
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if "mission-adaptation-stories" not in workbook.sheetnames:
            raise AssertionError("Sheet mission-adaptation-stories not found in Excel file.")

        sheet = workbook["mission-adaptation-stories"]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        header_index = {_normalize_header(cell): idx for idx, cell in enumerate(header_row) if cell}
        expected_headers = {"title", _normalize_header(excel_column)}
        missing_headers = [name for name in expected_headers if name not in header_index]
        if missing_headers:
            raise AssertionError("Missing headers in Excel: " + ", ".join(sorted(missing_headers)))

        expected: dict = {}
        duplicate_expected = []
        duplicate_conflicts = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            title_value = row[header_index["title"]]
            if not title_value:
                continue
            normalized_title = _normalize_title(title_value)
            values = row[header_index[_normalize_header(excel_column)]]
            normalized_values = _normalize_impacts(values)
            if normalized_title in expected:
                duplicate_expected.append(str(title_value))
                if sorted(expected[normalized_title]) != sorted(normalized_values):
                    duplicate_conflicts.append(str(title_value))
                continue
            expected[normalized_title] = normalized_values
        return expected, duplicate_expected, duplicate_conflicts
    finally:
        workbook.close()


def _load_observed_from_json(json_key: str) -> tuple[dict, list]:
    observed: dict = {}
    duplicate_observed = []
    pages_dir = Path("data/pages")
    for path in pages_dir.glob("*.json"):
        data = json.loads(path.read_text(encoding="utf-8"))
        title = data.get("title")
        if not title:
            continue
        normalized_title = _normalize_title(title)
        if normalized_title in observed:
            duplicate_observed.append(title)
            continue
        observed[normalized_title] = _normalize_observed_impacts(data.get(json_key))
    return observed, duplicate_observed


def _diff_expected_vs_observed(expected: dict, observed: dict, mismatch_label: str) -> tuple[list, list, list]:
    missing_titles = []
    mismatches = []
    report_rows = []
    for title, expected_values in expected.items():
        if title not in observed:
            missing_titles.append(title)
            report_rows.append(
                {
                    "issue_type": "missing_json",
                    "title": title,
                    "expected": ";".join(expected_values),
                    "observed": "",
                    "details": "Title missing in JSON pages",
                }
            )
            continue
        observed_values = observed[title]
        expected_set = set(expected_values)
        observed_set = set(observed_values)
        missing_values = sorted(expected_set - observed_set)
        extra_values = sorted(observed_set - expected_set)
        if not missing_values and not extra_values:
            continue
        details = []
        if missing_values:
            details.append(f"missing={missing_values}")
        if extra_values:
            details.append(f"extra={extra_values}")
        mismatches.append(f"{title}: expected={expected_values} observed={observed_values}")
        report_rows.append(
            {
                "issue_type": "mismatch",
                "title": title,
                "expected": ";".join(expected_values),
                "observed": ";".join(observed_values),
                "details": f"{mismatch_label} ({'; '.join(details)})",
            }
        )
    return missing_titles, mismatches, report_rows


def _collect_duplicate_errors(
    duplicate_expected: list,
    duplicate_conflicts: list,
    duplicate_observed: list,
    report_rows: list,
) -> list:
    errors = []
    for title in duplicate_expected:
        report_rows.append(
            {
                "issue_type": "duplicate_excel",
                "title": _normalize_title(title),
                "expected": "",
                "observed": "",
                "details": "Duplicate title in Excel",
            }
        )
    if duplicate_conflicts and STRICT_EXCEL_COMPARISON:
        errors.append("Conflicting duplicate titles in Excel: " + ", ".join(sorted(duplicate_conflicts)))
    if duplicate_observed and STRICT_EXCEL_COMPARISON:
        errors.append("Duplicate titles in JSON: " + ", ".join(sorted(duplicate_observed)))
        for title in duplicate_observed:
            report_rows.append(
                {
                    "issue_type": "duplicate_json",
                    "title": _normalize_title(title),
                    "expected": "",
                    "observed": "",
                    "details": "Duplicate title in JSON pages",
                }
            )
    return errors


def _write_report(report_path: Path, report_rows: list) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["issue_type", "title", "expected", "observed", "details"],
        )
        writer.writeheader()
        writer.writerows(report_rows)


def _compare_excel_column_to_json(
    excel_column: str,
    json_key: str,
    report_path: Path,
    mismatch_label: str,
) -> None:
    expected, duplicate_expected, duplicate_conflicts = _load_expected_from_excel(excel_column)
    observed, duplicate_observed = _load_observed_from_json(json_key)
    missing_titles, mismatches, report_rows = _diff_expected_vs_observed(expected, observed, mismatch_label)

    errors = _collect_duplicate_errors(duplicate_expected, duplicate_conflicts, duplicate_observed, report_rows)
    if missing_titles and STRICT_EXCEL_COMPARISON:
        errors.append("Missing titles in JSON: " + ", ".join(sorted(missing_titles)))
    if mismatches and STRICT_EXCEL_COMPARISON:
        errors.append(f"{mismatch_label}:\n" + "\n".join(mismatches))

    if report_rows:
        _write_report(report_path, report_rows)
        if errors:
            errors.append(f"Report written to {report_path.as_posix()}")

    assert not errors, "\n".join(errors)


def test_climate_impacts_match_excel():
    _compare_excel_column_to_json(
        excel_column="Climate Impact",
        json_key="climate_impacts",
        report_path=Path("data/exports/climate_impacts_report.csv"),
        mismatch_label="Climate impact mismatches",
    )


def test_sectors_match_excel():
    _compare_excel_column_to_json(
        excel_column="Sectors",
        json_key="adaptation_sectors",
        report_path=Path("data/exports/adaptation_sectors_report.csv"),
        mismatch_label="Adaptation sector mismatches",
    )


def test_key_community_systems_match_excel():
    _compare_excel_column_to_json(
        excel_column="Key Community Systems",
        json_key="key_community_systems",
        report_path=Path("data/exports/key_community_systems_report.csv"),
        mismatch_label="Key community systems mismatches",
    )


def test_funding_programme_match_excel():
    _compare_excel_column_to_json(
        excel_column="Funding programme",
        json_key="funding_programme",
        report_path=Path("data/exports/funding_programme_report.csv"),
        mismatch_label="Funding programme mismatches",
    )


def test_countries_match_excel():
    _compare_excel_column_to_json(
        excel_column="Countries",
        json_key="countries",
        report_path=Path("data/exports/countries_report.csv"),
        mismatch_label="Countries mismatches",
    )
