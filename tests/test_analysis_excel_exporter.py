import json
from pathlib import Path

import openpyxl

from exporters.analysis_excel_exporter import export_analysis_to_excel


def _write_sample_analysis(input_dir: Path) -> None:
    payload = {
        "url": "https://example.com/story",
        "title": "Example Story",
        "answers": {
            "Q1": "Line 1\nLine 2",
            "Q2": "Another answer",
        },
    }
    (input_dir / "sample.json").write_text(json.dumps(payload), encoding="utf-8")


def test_export_excel_applies_formatting_defaults(tmp_path: Path):
    input_dir = tmp_path / "analysis"
    input_dir.mkdir()
    output_path = tmp_path / "exports" / "analysis.xlsx"
    _write_sample_analysis(input_dir)

    export_analysis_to_excel(input_dir=input_dir, output_path=output_path, verbose=False)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.active
    assert sheet is not None

    assert sheet.freeze_panes == "A2"
    assert sheet["A1"].font.bold is True
    assert sheet["D2"].alignment.wrap_text is True
    assert sheet.column_dimensions["A"].width is not None


def test_export_excel_can_disable_formatting_options(tmp_path: Path):
    input_dir = tmp_path / "analysis"
    input_dir.mkdir()
    output_path = tmp_path / "exports" / "analysis.xlsx"
    _write_sample_analysis(input_dir)

    export_analysis_to_excel(
        input_dir=input_dir,
        output_path=output_path,
        verbose=False,
        header_bold=False,
        auto_width=False,
        wrap_text=False,
        freeze_panes=None,
    )

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.active
    assert sheet is not None

    assert sheet.freeze_panes is None
    assert sheet["A1"].font.bold is False
    assert sheet["D2"].alignment.wrap_text is not True
    assert sheet.column_dimensions["A"].width == 13.0


def test_export_excel_parses_answers_from_ai_result_when_answers_missing(tmp_path: Path):
    input_dir = tmp_path / "analysis"
    input_dir.mkdir()
    output_path = tmp_path / "exports" / "analysis.xlsx"
    payload = {
        "url": "https://example.com/story",
        "title": "Example Story",
        "answers": None,
        "ai_result": "Q1: First value\nQ2: Second value",
    }
    (input_dir / "sample.json").write_text(json.dumps(payload), encoding="utf-8")

    export_analysis_to_excel(input_dir=input_dir, output_path=output_path, verbose=False)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.active
    assert sheet is not None

    headers = [cell.value for cell in sheet[1]]
    assert "Answer 1" in headers
    assert "Answer 2" in headers
    answer_1_col = headers.index("Answer 1") + 1
    answer_2_col = headers.index("Answer 2") + 1
    assert sheet.cell(row=2, column=answer_1_col).value == "First value"
    assert sheet.cell(row=2, column=answer_2_col).value == "Second value"


def test_export_excel_uses_default_answer_columns_when_none_found(tmp_path: Path):
    input_dir = tmp_path / "analysis"
    input_dir.mkdir()
    output_path = tmp_path / "exports" / "analysis.xlsx"
    payload = {
        "url": "https://example.com/story",
        "title": "Example Story",
        "answers": None,
        "ai_result": "",
    }
    (input_dir / "sample.json").write_text(json.dumps(payload), encoding="utf-8")

    export_analysis_to_excel(input_dir=input_dir, output_path=output_path, verbose=False)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.active
    assert sheet is not None

    headers = [cell.value for cell in sheet[1]]
    assert headers[3] == "Answer 1"
    assert headers[-1] == "Answer 10"
